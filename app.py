from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash, Response, stream_with_context
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
from flask_mail import Mail, Message
import secrets
from flask_login import login_required
from openpyxl.styles import NamedStyle
from openpyxl import Workbook
import os
import re
import pandas as pd
import time
import json
from flask import jsonify

app = Flask(__name__)
app.secret_key = os.urandom(24)  # Clave secreta para producción

# ================= Configuración del servidor de correo =================

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'maria.victoriacr25@gmail.com'
app.config['MAIL_PASSWORD'] = 'jstc blpy psjc osil'
app.config['MAIL_DEFAULT_SENDER'] = ('Soporte Admin', 'tu_correo@gmail.com')

mail = Mail(app)

# ================= CONFIGURACIÓN DE ARCHIVOS =================

# Configuración de archivos
ARCHIVOS = {
    'consumos': 'data/consumos.xlsx',
    'historial': 'data/historial_pedidos.xlsx'
}

DATA_DIR = "data"
REPORTES_DIR = "reportes"
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(REPORTES_DIR, exist_ok=True)

EXCEL_ADMINS = os.path.join(DATA_DIR, "administradores.xlsx")
EXCEL_CLIENTES = os.path.join(DATA_DIR, "clientes.xlsx")
EXCEL_CONSUMOS = os.path.join(DATA_DIR, "consumos.xlsx")
EXCEL_PRODUCTOS = os.path.join(DATA_DIR, "productos.xlsx")
INVENTARIO_EXCEL = os.path.join(DATA_DIR, "inventario.xlsx")
EXCEL_MESAS = os.path.join(DATA_DIR, "mesas.xlsx")  # Nuevo archivo para mesas
VENTAS_ENTRADAS = os.path.join(DATA_DIR, "Ventas_Entradas.xlsx")

# Inicializar archivo de mesas si no existe
if not os.path.exists(EXCEL_MESAS):
    mesas_iniciales = [
        {"id": 1, "numero": 1, "nombre_cliente": ""},
        {"id": 2, "numero": 2, "nombre_cliente": ""},
        {"id": 3, "numero": 3, "nombre_cliente": ""},
        {"id": 4, "numero": 4, "nombre_cliente": ""},
        {"id": 5, "numero": 5, "nombre_cliente": ""},
        {"id": 6, "numero": 6, "nombre_cliente": ""},
        {"id": 7, "numero": 7, "nombre_cliente": ""},
        {"id": 8, "numero": 8, "nombre_cliente": ""},
        {"id": 9, "numero": 9, "nombre_cliente": ""},
        {"id": 10, "numero": 10, "nombre_cliente": ""}
    ]
    pd.DataFrame(mesas_iniciales).to_excel(EXCEL_MESAS, index=False)

# Función para cargar mesas
def cargar_mesas():
    return pd.read_excel(EXCEL_MESAS).to_dict("records")

# Función para obtener el nombre del archivo de ventas con la fecha actual
def obtener_nombre_archivo_ventas():
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    return os.path.join(DATA_DIR, f"ventas_{fecha_actual}.xlsx")

# Inicializar archivo de ventas si no existe
def inicializar_archivo_ventas():
    archivo_ventas = obtener_nombre_archivo_ventas()
    if not os.path.exists(archivo_ventas):
        pd.DataFrame(columns=["Tipo", "Fecha_Venta", "Producto", "Cantidad", "Total_Venta"]).to_excel(archivo_ventas, index=False)

inicializar_archivo_ventas()

# ================= INICIALIZACIÓN DE ARCHIVOS =================

# Inicializar archivo de ventas si no existe
if not os.path.exists(VENTAS_ENTRADAS):
    pd.DataFrame(columns=["Fecha_Venta", "Producto", "Cantidad", "Total_Venta"]).to_excel(VENTAS_ENTRADAS, index=False)

def inicializar_archivos():

    columnas_consumos = [
            "Mesa", 
            "Producto", 
            "Cantidad", 
            "Precio", 
            "Categoría", 
            "Fecha_Hora", 
            "Estado", 
            "Total"  # Añadir esta columna
        ]
        
    if not os.path.exists(ARCHIVOS['consumos']):
            pd.DataFrame(columns=columnas_consumos).to_excel(ARCHIVOS['consumos'], index=False)
    else:
        # Verificar y añadir columna Total si falta
        df = pd.read_excel(ARCHIVOS['consumos'])
        if 'Total' not in df.columns:
            df['Total'] = 0
            df.to_excel(ARCHIVOS['consumos'], index=False)

    if not os.path.exists(EXCEL_PRODUCTOS):
        # Crear con columnas correctas
        pd.DataFrame(columns=["Nombre", "Precio", "Categoría", "Existencias"]).to_excel(EXCEL_PRODUCTOS, index=False)
    else:
        # Verificar columnas
        df_productos = pd.read_excel(EXCEL_PRODUCTOS)
        if "Nombre" not in df_productos.columns:
            # Corregir si el archivo existente tiene "Producto" como columna
            df_productos.rename(columns={"Producto": "Nombre"}, inplace=True)
            df_productos.to_excel(EXCEL_PRODUCTOS, index=False)

    if not os.path.exists(INVENTARIO_EXCEL):
        pd.DataFrame(columns=["Producto", "Cantidad", "Costo_Unitario", "PVP", "Ganancia"]).to_excel(INVENTARIO_EXCEL, index=False)

    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
    if not os.path.exists(REPORTES_DIR):
        os.makedirs(REPORTES_DIR)
    
    # Administradores
    if not os.path.exists(EXCEL_ADMINS):
        # Creamos el Excel con las columnas Usuario, Correo y Clave
        pd.DataFrame(columns=["Usuario", "Correo", "Clave"]).to_excel(EXCEL_ADMINS, index=False)
    else:
        # Si ya existe, verificamos que tenga la columna Correo
        df_admins = pd.read_excel(EXCEL_ADMINS)
        cambios = False
        if "Correo" not in df_admins.columns:
            df_admins["Correo"] = ""
            cambios = True
        if "Token" not in df_admins.columns:
            df_admins["Token"] = ""
            cambios = True
        if "Token_Expira" not in df_admins.columns:
            df_admins["Token_Expira"] = ""
            cambios = True
        if cambios: 
            df_admins.to_excel(EXCEL_ADMINS, index=False)
    
    # Clientes
    if not os.path.exists(EXCEL_CLIENTES):
        pd.DataFrame(columns=["Cédula", "Nombre"]).to_excel(EXCEL_CLIENTES, index=False)
    
    # Productos
    if not os.path.exists(EXCEL_PRODUCTOS):
        productos_iniciales = [
            {"Nombre": "Salchipapas", "Precio": 3.50},
            {"Nombre": "Chochos con tostado", "Precio": 2.50},
            {"Nombre": "Coca cola pequeña", "Precio": 1.50},
            {"Nombre": "Gaseosa de sabores mediana", "Precio": 2.00},
            {"Nombre": "Agua sin gas", "Precio": 1.00},
            {"Nombre": "Fuze Tea mediano", "Precio": 2.20},
            {"Nombre": "Güitig grande", "Precio": 2.50},
            {"Nombre": "Coca cola grande", "Precio": 2.50},
            {"Nombre": "Gaseosa de sabores grande", "Precio": 2.50},
            {"Nombre": "Fuze Tea grande", "Precio": 2.50},
            {"Nombre": "Gatorade", "Precio": 2.80},
            {"Nombre": "Papas sin marca", "Precio": 1.00},
            {"Nombre": "Chifles de Sal", "Precio": 1.20},
            {"Nombre": "Chifles de Dulce", "Precio": 1.20},
            {"Nombre": "Galletitas", "Precio": 0.80}
        ]
        pd.DataFrame(productos_iniciales).to_excel(EXCEL_PRODUCTOS, index=False)
    else:
        # Siempre cargar el DataFrame primero
        df_productos = pd.read_excel(EXCEL_PRODUCTOS)
        
        # Verificar y corregir nombre de columna si es necesario
        if "Producto" in df_productos.columns:
            df_productos.rename(columns={"Producto": "Nombre"}, inplace=True)
        
        # Asegurar que todas las columnas necesarias existan
        columnas_requeridas = ["Nombre", "Precio", "Categoría", "Existencias"]
        cambios = False
        
        for col in columnas_requeridas:
            if col not in df_productos.columns:
                df_productos[col] = "" if col == "Categoría" else 0
                cambios = True
        
        if cambios:
            df_productos.to_excel(EXCEL_PRODUCTOS, index=False)
    
    # Consumos
    if not os.path.exists(EXCEL_CONSUMOS):
        df = pd.DataFrame(columns=["Mesa", "Producto", "Cantidad", "Precio", "Fecha_Hora", "Total"])
        df.to_excel(EXCEL_CONSUMOS, index=False)

inicializar_archivos()

# ================= DECORADORES DE AUTENTICACIÓN =================
def login_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logueado'):
            return redirect(url_for('login_admin'))
        return f(*args, **kwargs)
    return decorated_function

def login_caja_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('caja_logueada'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ================= RUTAS PRINCIPALES =================
@app.route("/", methods=["GET", "POST"])
def inicio():
    if request.method == "POST":
        rol = request.form.get("rol")
        if rol == "admin":
            return redirect(url_for('login_admin'))
        elif rol == "caja":
            return redirect(url_for('login'))
    return render_template("inicio.html")

# ================= SISTEMA ADMINISTRADOR =================
@app.route("/admin/login", methods=["GET", "POST"])
def login_admin():
    if request.method == "POST":
        usuario = request.form.get("usuario")
        clave = request.form.get("clave")
        
        df = pd.read_excel(EXCEL_ADMINS)
        admin = df[df["Usuario"] == usuario]
        
        if not admin.empty and check_password_hash(admin.iloc[0]["Clave"], clave):
            session['admin_logueado'] = True
            return redirect(url_for('panel_admin'))
        
        return render_template("login_admin.html", error="Credenciales incorrectas")
    return render_template("login_admin.html")

@app.route("/admin/registro", methods=["GET", "POST"])
def registro_admin():
    df = pd.read_excel(EXCEL_ADMINS)
    if len(df) >= 3:
        return render_template("error.html", mensaje="Máximo de 3 administradores alcanzado")

    if request.method == "POST":
        usuario = request.form.get("usuario")
        correo = request.form.get("correo")
        clave_plana = request.form.get("clave")

        # Validaciones básicas
        if not usuario or not correo or not clave_plana:
            return render_template("error.html", mensaje="Todos los campos son obligatorios")

        if "@" not in correo:
            return render_template("error.html", mensaje="Correo inválido")

        # Generar hash de la contraseña
        clave = generate_password_hash(clave_plana)

        # Crear nuevo registro
        nuevo_admin = pd.DataFrame([[usuario, correo, clave]], 
                                   columns=["Usuario", "Correo", "Clave"])
        df = pd.concat([df, nuevo_admin], ignore_index=True)
        df.to_excel(EXCEL_ADMINS, index=False)

        return redirect(url_for('login_admin'))

    return render_template("registro_admin.html")

@app.route("/admin/reset_password/<token>", methods=["GET", "POST"])
def reset_password_admin(token):
    df = pd.read_excel(EXCEL_ADMINS)
    # Buscar al admin con ese token
    admin = df[df["Token"] == token]

    if admin.empty:
        return render_template("error.html", mensaje="Token inválido o ya utilizado.")

    # Verificar expiración
    token_expira_str = admin.iloc[0]["Token_Expira"]
    token_expira = datetime.strptime(token_expira_str, "%Y-%m-%d %H:%M:%S")
    if datetime.now() > token_expira:
        return render_template("error.html", mensaje="El enlace de recuperación ha expirado.")

    if request.method == "POST":
        nueva_clave = request.form.get("nueva_clave")
        confirma_clave = request.form.get("confirma_clave")

        if not nueva_clave or not confirma_clave:
            return render_template("reset_password_admin.html", 
                                   error="Debe llenar ambos campos.", token=token)

        if nueva_clave != confirma_clave:
            return render_template("reset_password_admin.html", 
                                   error="Las contraseñas no coinciden.", token=token)

        # Actualizar la contraseña en Excel
        index_admin = admin.index[0]
        df.at[index_admin, "Clave"] = generate_password_hash(nueva_clave)
        # Limpiar el token para que no se pueda reutilizar
        df.at[index_admin, "Token"] = ""
        df.at[index_admin, "Token_Expira"] = ""
        df.to_excel(EXCEL_ADMINS, index=False)

        return render_template("reset_password_admin.html", 
                               mensaje="Tu contraseña ha sido restablecida. ¡Puedes iniciar sesión!")

    return render_template("reset_password_admin.html", token=token)

@app.route('/admin/panel')
@login_admin_required
def panel_admin():
    return render_template("panel_admin.html")

@app.route("/admin/productos", methods=["GET", "POST"])
@login_admin_required
def admin_productos():
    df = pd.read_excel(EXCEL_PRODUCTOS)
    
    if request.method == "POST":
        if 'eliminar' in request.form:
            indice = int(request.form.get('indice'))
            df = df.drop(index=indice).reset_index(drop=True)
            df.to_excel(EXCEL_PRODUCTOS, index=False)
            return redirect(url_for('admin_productos'))
        
        if 'guardar' in request.form:
            indice = request.form.get('indice')
            producto = request.form.get('producto')
            precio = float(request.form.get('precio'))
            existencias = int(request.form.get('existencias'))
            categoria = request.form.get('categoria')  # Nueva categoría
            
            if indice and indice != 'None':
                indice = int(indice)
                df.at[indice, 'Nombre'] = producto
                df.at[indice, 'Precio'] = precio
                df.at[indice, 'Existencias'] = existencias
                df.at[indice, 'Categoría'] = categoria
            else:
                nuevo = pd.DataFrame([[producto, precio, categoria, existencias]], 
                                   columns=["Nombre", "Precio", "Categoría", "Existencias"])
                df = pd.concat([df, nuevo], ignore_index=True)
            
            df.to_excel(EXCEL_PRODUCTOS, index=False)
            return redirect(url_for('admin_productos'))
    
    return render_template("admin_productos.html", 
                         productos=df.to_dict("records"), 
                         enumerate=enumerate)

# Añadir esto temporalmente en tu código para actualizar archivos existentes
def corregir_consumos():
    try:
        df = pd.read_excel(ARCHIVOS['consumos'])
        if 'Total' not in df.columns:
            df['Total'] = df['Cantidad'] * df['Precio']
            df.to_excel(ARCHIVOS['consumos'], index=False)
    except Exception as e:
        print(f"Error al corregir consumos: {str(e)}")

corregir_consumos()

@app.route("/admin/clientes", methods=["GET", "POST"])
@login_admin_required
def admin_clientes():
    df = pd.read_excel(EXCEL_CLIENTES)
    
    if request.method == "POST":
        if 'eliminar' in request.form:
            indice = int(request.form.get('indice'))
            df = df.drop(index=indice).reset_index(drop=True)
            df.to_excel(EXCEL_CLIENTES, index=False)
            return redirect(url_for('admin_clientes'))
        
        if 'guardar' in request.form:
            indice = request.form.get('indice')
            cedula = request.form.get('cedula')
            nombre = request.form.get('nombre')
            
            if indice and indice != 'None':
                indice = int(indice)
                df.at[indice, 'Cédula'] = str(cedula)
                df.at[indice, 'Nombre'] = nombre
            else:
                if str(cedula) in df['Cédula'].astype(str).values:
                    return render_template("error.html", mensaje="La cédula ya está registrada")
                nuevo = pd.DataFrame([[str(cedula), nombre]], columns=["Cédula", "Nombre"])
                df = pd.concat([df, nuevo], ignore_index=True)
            
            df.to_excel(EXCEL_CLIENTES, index=False)
            return redirect(url_for('admin_clientes'))
    
    return render_template("admin_clientes.html", 
                         clientes=df.to_dict("records"), 
                         enumerate=enumerate)

# ================= RUTA PARA GENERAR REPORTES =================
@app.route("/admin/reportes", methods=["GET", "POST"])
@login_admin_required
def generar_reportes():
    if request.method == "POST":
        try:
            # Obtener la fecha del formulario
            fecha = request.form.get("fecha")
            fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")  # Convertir a objeto datetime
            fecha_str = fecha_obj.strftime("%Y-%m-%d")  # Formatear como YYYY-MM-DD

            # Leer archivos Excel
            archivo_ventas = obtener_nombre_archivo_ventas()
            df_ventas = pd.read_excel(archivo_ventas)

            # Filtrar ventas por fecha
            ventas_filtradas = df_ventas[
                df_ventas["Fecha_Venta"].str.contains(fecha_str)
            ]

            # Crear el archivo de reporte
            reporte_path = os.path.join(REPORTES_DIR, f"reporte_ventas_{fecha_str}.xlsx")

            # Crear un nuevo archivo Excel con openpyxl
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Ventas"

            # Escribir los encabezados
            for col_num, header in enumerate(ventas_filtradas.columns, 1):
                worksheet.cell(row=1, column=col_num, value=header)

            # Escribir los datos
            for row_num, row_data in enumerate(ventas_filtradas.values, 2):
                for col_num, cell_data in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col_num, value=cell_data)

            # Aplicar formato de moneda a la columna "Total_Venta"
            dinero_style = NamedStyle(name="dinero_style", number_format='"$"#,##0.00')
            for cell in worksheet["E"]:  # Columna E es "Total_Venta"
                cell.style = dinero_style

            # Guardar el archivo Excel
            workbook.save(reporte_path)

            # Enviar el archivo como descarga
            return send_file(reporte_path, as_attachment=True)
        
        except Exception as e:
            return render_template("error.html", mensaje=f"Error al generar el reporte: {str(e)}")
    
    return render_template("generar_reporte.html")

# ================= SISTEMA CAJA =================
@app.route("/caja/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        session['caja_logueada'] = True
        seccion = request.form.get("seccion")
        
        if seccion == "bar":
            return redirect(url_for('mesas'))
        elif seccion == "entradas":
            return redirect(url_for('venta_entradas'))  # Nueva función a crear
        elif seccion == "cocina":
            return redirect(url_for('cocina'))
    
    return render_template("login.html")

@app.route("/caja/mesas")
@login_caja_required
def mesas():
    mesas = cargar_mesas()
    return render_template("clientes.html", mesas=mesas)

@app.route("/caja/actualizar_mesa/<int:mesa_id>", methods=["POST"])
@login_caja_required
def actualizar_mesa(mesa_id):
    nombre_cliente = request.form.get("nombre_cliente")
    
    df_mesas = pd.read_excel(EXCEL_MESAS)
    df_mesas.loc[df_mesas["id"] == mesa_id, "nombre_cliente"] = nombre_cliente
    df_mesas.to_excel(EXCEL_MESAS, index=False)
    
    flash("Nombre del cliente actualizado correctamente", "success")
    return redirect(url_for('mesas'))

# ================= RUTA PARA REGISTRAR CONSUMOS =================
@app.route("/caja/registrar_consumo/<int:mesa>", methods=["GET", "POST"])
@login_caja_required
def registrar_consumo(mesa):
    # Cargar los productos desde el archivo Excel
    df_productos = pd.read_excel(EXCEL_PRODUCTOS)
    productos = df_productos.to_dict("records")

    if request.method == "POST":
        # Obtener los datos del formulario
        producto_id = int(request.form.get("producto"))
        cantidad = int(request.form.get("cantidad", 0))

        # Obtener el producto seleccionado
        producto_seleccionado = df_productos.iloc[producto_id]
        categoria = producto_seleccionado["Categoría"]  # Asegurar que se guarde la categoría
        nombre_producto = producto_seleccionado["Nombre"]
        precio_unitario = producto_seleccionado["Precio"]
        total = precio_unitario * cantidad

        # Registrar el consumo en el archivo Excel
        fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df_consumos = pd.read_excel(EXCEL_CONSUMOS)
        if "Categoría" not in df_consumos.columns:
            df_consumos["Categoría"] = "Comida"  # O actualizar con los valores reales
            df_consumos.to_excel(EXCEL_CONSUMOS, index=False)

        # Verificar si el producto ya está registrado para esta mesa
        consumo_existente = df_consumos[(df_consumos["Mesa"] == mesa) & (df_consumos["Producto"] == nombre_producto)]
        if not consumo_existente.empty:
            # Si el producto ya está registrado, sumar la cantidad
            indice = consumo_existente.index[0]
            df_consumos.at[indice, "Cantidad"] += cantidad
            df_consumos.at[indice, "Total"] += total
        else:
            # Si el producto no está registrado, agregar un nuevo consumo
            nuevo_consumo = pd.DataFrame([{
                "Mesa": mesa,
                "Producto": nombre_producto,
                "Cantidad": cantidad,
                "Precio": precio_unitario,
                "Categoría": categoria,
                "Fecha_Hora": fecha_hora,
                "Estado": "Pendiente",
                "Total": cantidad * precio_unitario  # Calcular el total
            }])
            df_consumos = pd.concat([df_consumos, nuevo_consumo], ignore_index=True)

        df_consumos.to_excel(EXCEL_CONSUMOS, index=False)
        return redirect(url_for('registrar_consumo', mesa=mesa))

    # Obtener los consumos registrados para esta mesa
    df_consumos = pd.read_excel(ARCHIVOS['consumos'])
    consumos_mesa = df_consumos[df_consumos["Mesa"] == mesa].to_dict("records")

    # Calcular total de manera segura
    total_pagar = df_consumos[df_consumos["Mesa"] == mesa]["Total"].sum()
    
    return render_template("registrar_consumo.html",
                         mesa=mesa,
                         productos=productos,
                         consumos=consumos_mesa,
                         total_pagar=total_pagar)

# ================= RUTA PARA PROCESAR EL PAGO =================
@app.route("/caja/procesar_pago/<int:mesa>", methods=["POST"])
@login_caja_required
def procesar_pago(mesa):
    metodo_pago = request.form.get("metodo_pago")
    referencia = request.form.get("referencia", "")

    # Obtener los consumos de la mesa
    df_consumos = pd.read_excel(EXCEL_CONSUMOS)
    consumos_mesa = df_consumos[df_consumos["Mesa"] == mesa]

    # Registrar los consumos pagados en el archivo de ventas
    fecha_venta = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    archivo_ventas = obtener_nombre_archivo_ventas()
    df_ventas = pd.read_excel(archivo_ventas)

    for _, consumo in consumos_mesa.iterrows():
        nueva_venta = {
            "Tipo": "Consumo",
            "Fecha_Venta": fecha_venta,
            "Producto": consumo["Producto"],
            "Cantidad": consumo["Cantidad"],
            "Total_Venta": consumo["Total"],
            "Detalle": ""  # Los consumos no tienen detalle
        }
        df_ventas = pd.concat([df_ventas, pd.DataFrame([nueva_venta])], ignore_index=True)

    df_ventas.to_excel(archivo_ventas, index=False)

    # Marcar los consumos de la mesa como pagados
    df_consumos.loc[df_consumos["Mesa"] == mesa, "Estado"] = "Pagado"
    df_consumos.loc[df_consumos["Mesa"] == mesa, "Metodo_Pago"] = metodo_pago
    if metodo_pago == "Transferencia":
        df_consumos.loc[df_consumos["Mesa"] == mesa, "Referencia"] = referencia
    df_consumos.to_excel(EXCEL_CONSUMOS, index=False)

    # Restablecer el nombre del cliente en la mesa
    df_mesas = pd.read_excel(EXCEL_MESAS)
    df_mesas.loc[df_mesas["numero"] == mesa, "nombre_cliente"] = ""
    df_mesas.to_excel(EXCEL_MESAS, index=False)

    # Eliminar los consumos de la mesa (liberar la mesa)
    df_consumos = df_consumos[df_consumos["Mesa"] != mesa]  # Filtrar y eliminar consumos de la mesa
    df_consumos.to_excel(EXCEL_CONSUMOS, index=False)

    # Mostrar mensaje de éxito
    flash("Pago exitoso. La mesa ha sido liberada.", "success")
    return redirect(url_for('registrar_consumo', mesa=mesa))

# ================= RUTA PARA REGISTRAR VENTA DE ENTRADAS =================
@app.route('/venta_entradas', methods=['GET', 'POST'])
def venta_entradas():
    if request.method == "POST":
        fecha = request.form.get("fecha")
        adultos = int(request.form.get("adultos", 0))
        ninos = int(request.form.get("ninos", 0))
        tercera_edad = int(request.form.get("tercera_edad", 0))
        detalle = request.form.get("detalle", "").strip()

        total = (adultos * 6) + (ninos * 4) + (tercera_edad * 4)

        # Registrar la venta de entradas en el archivo de ventas
        fecha_venta = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        archivo_ventas = obtener_nombre_archivo_ventas()
        df_ventas = pd.read_excel(archivo_ventas)

        nueva_venta = {
            "Tipo": "Entrada",
            "Fecha_Venta": fecha_venta,
            "Producto": "Entradas",
            "Cantidad": adultos + ninos + tercera_edad,
            "Total_Venta": total,
            "Detalle": detalle  # Incluir el detalle de la venta de entradas
        }
        df_ventas = pd.concat([df_ventas, pd.DataFrame([nueva_venta])], ignore_index=True)
        df_ventas.to_excel(archivo_ventas, index=False)

        # Guardar la venta de entradas en el archivo específico
        datos = {
            "Fecha": [fecha],
            "Hora": datetime.now().strftime("%H:%M:%S"),
            "Adultos": [adultos],
            "Niños": [ninos],
            "Tercera_Edad": [tercera_edad],
            "Total_Venta": [total],
            "Detalle": [detalle]
        }
        df_nueva_venta = pd.DataFrame(datos)

        if os.path.exists(VENTAS_ENTRADAS):
            df_existente = pd.read_excel(VENTAS_ENTRADAS)
            df_final = pd.concat([df_existente, df_nueva_venta], ignore_index=True)
        else:
            df_final = df_nueva_venta

        df_final.to_excel(VENTAS_ENTRADAS, index=False)

        return redirect(url_for('venta_entradas', success=1))

    success = request.args.get('success', 0)
    return render_template("venta_entradas.html", success=success)

@app.route("/caja/cerrar_caja")
@login_caja_required
def cerrar_caja():
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    
    df = pd.read_excel(EXCEL_CONSUMOS)
    try:
        consumos_hoy = df[df["Fecha_Hora"].str.contains(fecha_actual)]
    except:
        consumos_hoy = pd.DataFrame(columns=CONSUMOS_COLS)
    
    resumen_general = {
        "total_cancelado": 0,
        "total_pendiente": 0,
        "detalle": []
    }
    
    clientes_unico = df["Cédula"].unique()
    resumen_clientes = []
    
    for cedula in clientes_unico:
        cliente_consumos = df[(df["Cédula"] == cedula) & 
                             (df["Fecha_Hora"].str.contains(fecha_actual))]
        
        total_cliente = sum(row["Cantidad"] * row["Precio"] 
                           for _, row in cliente_consumos.iterrows())
        
        estado = "CANCELADO" if all(cliente_consumos["Estado"] == "CANCELADO") else "PENDIENTE"
        
        resumen_clientes.append({
            "cedula": cedula,
            "total": total_cliente,
            "estado": estado
        })
        
        if estado == "CANCELADO":
            resumen_general["total_cancelado"] += total_cliente
        else:
            resumen_general["total_pendiente"] += total_cliente
    
    resumen_general["total_general"] = resumen_general["total_cancelado"] + resumen_general["total_pendiente"]
    
    return render_template("cierre_caja.html",
                         fecha=fecha_actual,
                         resumen_clientes=resumen_clientes,
                         resumen_general=resumen_general)

# ================= RUTA PARA COCINA =================

# Ruta para despachar pedidos
@app.route('/despachar_pedido/<int:pedido_id>', methods=['DELETE'])
def despachar_pedido(pedido_id):
    try:
        df_consumos = pd.read_excel(ARCHIVOS['consumos'])
        pedido = df_consumos.iloc[pedido_id].to_dict()
        
        # Registrar en historial
        df_historial = pd.read_excel(ARCHIVOS['historial']) if os.path.exists(ARCHIVOS['historial']) else pd.DataFrame()
        
        nuevo_registro = {
            "Mesa": pedido['Mesa'],
            "Producto": pedido['Producto'],
            "Cantidad": pedido['Cantidad'],
            "Fecha_Hora": pedido['Fecha_Hora'],
            "Fecha_Despacho": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Tiempo_Preparacion": (datetime.now() - datetime.strptime(pedido['Fecha_Hora'], "%Y-%m-%d %H:%M:%S")).seconds // 60
        }
        
        df_historial = pd.concat([df_historial, pd.DataFrame([nuevo_registro])], ignore_index=True)
        df_historial.to_excel(ARCHIVOS['historial'], index=False)
        
        # Eliminar de consumos
        df_consumos = df_consumos.drop(pedido_id)
        df_consumos.to_excel(ARCHIVOS['consumos'], index=False)
        
        return jsonify({
            'success': True,
            'mesa': pedido['Mesa'],
            'producto': pedido['Producto']
        })
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

# Ruta para detalles del pedido
@app.route('/detalles_pedido/<int:pedido_id>')
def detalles_pedido(pedido_id):
    try:
        df = pd.read_excel(ARCHIVOS['consumos'])
        pedido = df.iloc[pedido_id].to_dict()
        return f"""
            <p><strong>Producto:</strong> {pedido['Producto']}</p>
            <p><strong>Cantidad:</strong> {pedido['Cantidad']}</p>
            <p><strong>Mesa:</strong> {pedido['Mesa']}</p>
            <p><strong>Hora Pedido:</strong> {pedido['Fecha_Hora']}</p>
            <p><strong>Notas:</strong> {pedido.get('Notas', 'N/A')}</p>
        """
    except:
        return "Error al cargar detalles"
    
# Ruta principal de cocina
@app.route("/cocina")
def cocina():
    try:
        # Cargar y filtrar consumos
        df_consumos = pd.read_excel(ARCHIVOS['consumos'])
        pedidos_comida = df_consumos[
            (df_consumos["Categoría"] == "Comida") & 
            (df_consumos["Estado"] == "Pendiente")
        ].reset_index(drop=True)

        # Procesar tiempos y preparar datos para template
        ahora = datetime.now()
        pedidos_procesados = []
        
        for idx, row in pedidos_comida.iterrows():
            fecha_pedido = datetime.strptime(row['Fecha_Hora'], "%Y-%m-%d %H:%M:%S")
            tiempo_espera = (ahora - fecha_pedido).seconds // 60  # En minutos
            
            pedidos_procesados.append({
                'id': idx,
                'producto': row['Producto'],
                'cantidad': row['Cantidad'],
                'mesa': row['Mesa'],
                'fecha_hora': row['Fecha_Hora'],
                'tiempo_espera': tiempo_espera,
                'timestamp': int(fecha_pedido.timestamp()),
                'categoria': row['Categoría']
            })

        # Calcular tiempo promedio de preparación
        promedio_preparacion = 0
        if os.path.exists(ARCHIVOS['historial']):
            df_historial = pd.read_excel(ARCHIVOS['historial'])
            if not df_historial.empty and 'Tiempo_Preparacion' in df_historial.columns:
                promedio_preparacion = df_historial['Tiempo_Preparacion'].mean().round(1)

        return render_template(
            "cocina.html",
            pedidos=pedidos_procesados,
            promedio_preparacion=promedio_preparacion,
            sonido_activo=True
        )
        
    except Exception as e:
        print(f"Error en cocina: {str(e)}")
        return render_template("error.html", mensaje="Error al cargar los pedidos de cocina")

# ================= RUTA PARA MARCAR PEDIDO COMO LISTO =================
@app.route("/cocina/completar_pedido/<int:pedido_id>", methods=["POST"])
def completar_pedido(pedido_id):
    df_consumos = pd.read_excel(EXCEL_CONSUMOS).reset_index(drop=True)  # Reiniciar índice
    
    if pedido_id in df_consumos.index:
        df_consumos.at[pedido_id, "Estado"] = "Completado"
        df_consumos.to_excel(EXCEL_CONSUMOS, index=False)
        
        mesa = df_consumos.at[pedido_id, "Mesa"]
        print(f"Pedido {pedido_id} completado para la mesa {mesa}")  # Log para depuración
        flash("Pedido marcado como completado", "success")
    else:
        flash("Pedido no encontrado", "error")
    
    return redirect(url_for('cocina'))

@app.route("/admin/inventario", methods=["GET", "POST"])
@login_admin_required
def admin_inventario():
    df_inv = pd.read_excel(INVENTARIO_EXCEL)
    df_prod = pd.read_excel(EXCEL_PRODUCTOS)

    if request.method == "POST":
        try:
            # Acción: Crear nuevo producto
            if 'guardar' in request.form:
                producto = request.form.get("producto")
                cantidad = int(request.form.get("cantidad"))
                costo_unitario = float(request.form.get("costo_unitario"))
                pvp = float(request.form.get("pvp"))
                categoria = request.form.get("categoria")  # Nuevo campo: Categoría
                ganancia = round(pvp - costo_unitario, 2)

                # Verificar si el producto ya existe en el inventario
                if producto in df_inv["Producto"].values:
                    flash("El producto ya existe en el inventario.", "error")
                    return redirect(url_for('admin_inventario'))

                # Crear nuevo registro en el inventario
                nuevo_inv = pd.DataFrame([{
                    "Producto": producto,
                    "Cantidad": cantidad,
                    "Costo_Unitario": costo_unitario,
                    "PVP": pvp,
                    "Ganancia": ganancia,
                    "Categoría": categoria  # Nueva columna: Categoría
                }])
                df_inv = pd.concat([df_inv, nuevo_inv], ignore_index=True)
                df_inv.to_excel(INVENTARIO_EXCEL, index=False)

                # Actualizar productos
                if producto in df_prod["Nombre"].values:
                    df_prod.loc[df_prod["Nombre"] == producto, "Existencias"] += cantidad
                    df_prod.loc[df_prod["Nombre"] == producto, "Categoría"] = categoria
                else:
                    nuevo_prod = pd.DataFrame([{
                        "Nombre": producto,
                        "Precio": pvp,
                        "Categoría": categoria,  # Nueva columna: Categoría
                        "Existencias": cantidad
                    }])
                    df_prod = pd.concat([df_prod, nuevo_prod], ignore_index=True)
                df_prod.to_excel(EXCEL_PRODUCTOS, index=False)

                flash("Producto agregado correctamente.", "success")
                return redirect(url_for('admin_inventario'))

            # Acción: Eliminar producto
            elif 'eliminar' in request.form:
                indice = request.form.get("indice")
                if not indice:
                    flash("Índice no válido.", "error")
                    return redirect(url_for('admin_inventario'))

                indice = int(indice)
                
                # Verificar si el índice es válido
                if indice not in df_inv.index:
                    flash("Índice no válido.", "error")
                    return redirect(url_for('admin_inventario'))

                # Obtener el nombre del producto a eliminar
                producto_eliminado = df_inv.at[indice, "Producto"]

                # Eliminar el producto del inventario
                df_inv = df_inv.drop(index=indice).reset_index(drop=True)
                df_inv.to_excel(INVENTARIO_EXCEL, index=False)

                # Eliminar el producto de la lista de productos (opcional)
                df_prod = df_prod[df_prod["Nombre"] != producto_eliminado]
                df_prod.to_excel(EXCEL_PRODUCTOS, index=False)

                flash("Producto eliminado correctamente.", "success")
                return redirect(url_for('admin_inventario'))

            # Acción: Editar producto
            elif 'editar' in request.form:
                indice = request.form.get("indice")
                if not indice:
                    flash("Índice no válido.", "error")
                    return redirect(url_for('admin_inventario'))

                indice = int(indice)
                producto = request.form.get("producto")
                cantidad = int(request.form.get("cantidad"))
                costo_unitario = float(request.form.get("costo_unitario"))
                pvp = float(request.form.get("pvp"))
                categoria = request.form.get("categoria")  # Nuevo campo: Categoría
                ganancia = round(pvp - costo_unitario, 2)

                # Verificar si el índice es válido
                if indice not in df_inv.index:
                    flash("Índice no válido.", "error")
                    return redirect(url_for('admin_inventario'))

                # Actualizar valores en el inventario
                df_inv.at[indice, "Producto"] = producto
                df_inv.at[indice, "Cantidad"] = cantidad
                df_inv.at[indice, "Costo_Unitario"] = costo_unitario
                df_inv.at[indice, "PVP"] = pvp
                df_inv.at[indice, "Ganancia"] = ganancia
                df_inv.at[indice, "Categoría"] = categoria  # Nueva columna: Categoría
                df_inv.to_excel(INVENTARIO_EXCEL, index=False)

                # Actualizar productos
                if producto in df_prod["Nombre"].values:
                    df_prod.loc[df_prod["Nombre"] == producto, "Precio"] = pvp
                    df_prod.loc[df_prod["Nombre"] == producto, "Existencias"] = cantidad
                    df_prod.loc[df_prod["Nombre"] == producto, "Categoría"] = categoria
                else:
                    flash("Producto no encontrado en la lista de productos.", "error")
                    return redirect(url_for('admin_inventario'))

                df_prod.to_excel(EXCEL_PRODUCTOS, index=False)

                flash("Producto actualizado correctamente.", "success")
                return redirect(url_for('admin_inventario'))

        except Exception as e:
            flash(f"Error: {str(e)}", "error")
            return redirect(url_for('admin_inventario'))

    # Convertir el DataFrame a una lista de diccionarios para la plantilla
    inventario = df_inv.to_dict("records")
    return render_template("admin_inventario.html", inventario=inventario, enumerate=enumerate)

@app.route("/admin/mesas", methods=["GET", "POST"])
@login_admin_required
def admin_mesas():
    df_mesas = pd.read_excel(EXCEL_MESAS)
    
    if request.method == "POST":
        if 'agregar' in request.form:
            # Agregar una nueva mesa
            nueva_mesa = {
                "id": df_mesas["id"].max() + 1 if not df_mesas.empty else 1,
                "numero": int(request.form.get("numero")),
                "nombre_cliente": ""
            }
            df_mesas = pd.concat([df_mesas, pd.DataFrame([nueva_mesa])], ignore_index=True)
            df_mesas.to_excel(EXCEL_MESAS, index=False)
            return redirect(url_for('admin_mesas'))
        
        if 'eliminar' in request.form:
            # Eliminar una mesa
            mesa_id = int(request.form.get("mesa_id"))
            df_mesas = df_mesas[df_mesas["id"] != mesa_id]
            df_mesas.to_excel(EXCEL_MESAS, index=False)
            return redirect(url_for('admin_mesas'))
    
    return render_template("admin_mesas.html", mesas=df_mesas.to_dict("records"))

@app.route("/admin/actualizar_existencias/<int:producto_id>", methods=["POST"])
@login_admin_required
def actualizar_existencias(producto_id):
    df_productos = pd.read_excel(EXCEL_PRODUCTOS)
    
    if producto_id in df_productos.index:
        nueva_existencia = int(request.form.get("existencias"))
        df_productos.at[producto_id, "Existencias"] = nueva_existencia
        df_productos.to_excel(EXCEL_PRODUCTOS, index=False)
        flash("Existencias actualizadas correctamente", "success")
    else:
        flash("Producto no encontrado", "error")
    
    return redirect(url_for('admin_productos'))

@app.route("/admin/reportes_rango", methods=["GET", "POST"])
@login_admin_required
def generar_reportes_rango():
    if request.method == "POST":
        fecha_inicio = request.form.get("fecha_inicio")
        fecha_fin = request.form.get("fecha_fin")
        
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_obj = datetime.strptime(fecha_fin, "%Y-%m-%d")
            
            archivo_ventas = obtener_nombre_archivo_ventas()
            df_ventas = pd.read_excel(archivo_ventas)
            
            df_ventas["Fecha_Venta"] = pd.to_datetime(df_ventas["Fecha_Venta"])
            ventas_filtradas = df_ventas[
                (df_ventas["Fecha_Venta"] >= fecha_inicio_obj) &
                (df_ventas["Fecha_Venta"] <= fecha_fin_obj)
            ]
            
            reporte_path = os.path.join(REPORTES_DIR, f"reporte_ventas_{fecha_inicio}_a_{fecha_fin}.xlsx")
            ventas_filtradas.to_excel(reporte_path, index=False)
            
            return send_file(reporte_path, as_attachment=True)
        
        except Exception as e:
            return render_template("error.html", mensaje=f"Error al generar el reporte: {str(e)}")
    
    return render_template("generar_reporte_rango.html")

@app.route("/admin/actualizar_inventario/<int:producto_id>", methods=["POST"])
@login_admin_required
def actualizar_inventario(producto_id):
    df_inventario = pd.read_excel(INVENTARIO_EXCEL)
    
    if producto_id in df_inventario.index:
        nueva_cantidad = int(request.form.get("cantidad"))
        df_inventario.at[producto_id, "Cantidad"] = nueva_cantidad
        df_inventario.to_excel(INVENTARIO_EXCEL, index=False)
        flash("Inventario actualizado correctamente", "success")
    else:
        flash("Producto no encontrado", "error")
    
    return redirect(url_for('admin_inventario'))

@app.route("/admin/historial_entradas")
@login_admin_required
def historial_entradas():
    df_ventas = pd.read_excel(VENTAS_ENTRADAS)
    return render_template("historial_entradas.html", ventas=df_ventas.to_dict("records"))

@app.route("/admin/historial_consumos/<int:mesa>")
@login_admin_required
def historial_consumos(mesa):
    df_consumos = pd.read_excel(EXCEL_CONSUMOS)
    consumos_mesa = df_consumos[df_consumos["Mesa"] == mesa].to_dict("records")
    return render_template("historial_consumos.html", consumos=consumos_mesa, mesa=mesa)

@app.route("/admin/recuperar_password", methods=["GET", "POST"])
def recuperar_password_admin():
    if request.method == "POST":
        email = request.form.get("email")
        
        df = pd.read_excel(EXCEL_ADMINS)
        # Buscar admin por correo
        admin = df[df["Correo"] == email]

        if admin.empty:
            return render_template("recuperar_password_admin.html", 
                                   error="No se encontró un administrador con ese correo.")

        # Generar token aleatorio
        token = secrets.token_urlsafe(32)

        # Definir fecha/hora de expiración (ej. 30 minutos)
        expira = datetime.now() + timedelta(minutes=30)
        expira_str = expira.strftime("%Y-%m-%d %H:%M:%S")

        # Actualizar el registro en Excel
        index_admin = admin.index[0]
        df.at[index_admin, "Token"] = token
        df.at[index_admin, "Token_Expira"] = expira_str
        df.to_excel(EXCEL_ADMINS, index=False)

        # Enviar correo con Flask-Mail
        reset_url = url_for('reset_password_admin', token=token, _external=True)
        msg = Message("Complejo Recreacional Maria Victoria - Recuperación de Contraseña",
                      recipients=[email])
        msg.body = f"""Hola,

Has solicitado restablecer tu contraseña. Haz clic en el siguiente enlace (o pégalo en tu navegador) para continuar:

{reset_url}

Este enlace expirará en 30 minutos.

Si no solicitaste restablecer tu contraseña, ignora este correo.
"""
        mail.send(msg)

        return render_template("recuperar_password_admin.html", 
                               mensaje="Se envió un enlace de recuperación a tu correo.")
    
    return render_template("recuperar_password_admin.html")

# ================= FUNCIONES GENERALES =================
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('inicio'))

if __name__ == "__main__":
    app.run(debug=True)  # Cambiar a app.run en lugar de socketio.run